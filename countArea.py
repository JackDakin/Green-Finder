from Modules import detectGreen as dg
import os
from PIL import Image
from openpyxl import Workbook

src_dir = r'C:\Users\jackf\Pictures\PFR PLANT CODING\Additional Work 19.07\3.Mask'
dst_dir = r'C:\Users\jackf\Documents\PFR WORK\PLANT CODING\Additional Work 19.07'

xl_file = os.path.join(dst_dir, 'additional_plant_data.xlsx')

wb = Workbook()
ws = wb.active



ws['A1'] = "Location"
ws['A2'] = "Position"
ws['A4'] = "Image size"

#ws.cell(row=xl_row, column=xl_col, value="Value")

'''
# get the part of the file names that are shared
def get_names(names_list):

    for name in names_list:

        # remove all junk around the name
        # i.e. "IMG_1407 BJPP_22_0004.9 side" should become "BJPP_22_0004.9"

    return
'''


xl_row = 1
xl_col = 3
test = 0

for current_dir, dirnames, filenames in os.walk(src_dir):
    test += 1
    print("\n", test)
    print(f"{current_dir}\n{dirnames}\n{filenames}")


    # if dir has files in it
    if not filenames:
        pass
    else:
        xl_row = 1
        # save the current directory to the excel file
        print("files present")
        location = current_dir.replace(r"C:\Users\jackf\Pictures\PFR PLANT CODING\crop_masks", "")
        ws.cell(row=xl_row, column=xl_col, value=location)

        # get the size of the first image in this folder
        xl_row = 4
        im = Image.open(os.path.join(current_dir, filenames[0]))
        w, h = im.size
        size = w * h
        # save this to the excel file
        ws.cell(row=xl_row, column=xl_col, value=size)

        im.close()



        xl_row = 6
        for file in filenames:
            print(file)

            # save the name of the current image to excel file
            pix_count = dg.countGreen(file, current_dir)

            #save the name of the file to the left of the number
            name_col = xl_col - 1
            ws.cell(row=xl_row, column=name_col, value=file)


            # save the white pixel count to the excel file
            ws.cell(row=xl_row, column=xl_col, value=pix_count)
            xl_row += 1

        xl_col += 3


wb.save(filename=xl_file)
print("\nSaved to Excel")
